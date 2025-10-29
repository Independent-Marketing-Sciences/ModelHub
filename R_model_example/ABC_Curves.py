import sys
import warnings
import os

warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

try:
    import pandas as pd
    import numpy as np
    from scipy import optimize
    from openpyxl import load_workbook
    from imsciences import *
except ImportError as e:
    print(f"Missing required package: {e}")
    print("\nInstall all requirements with:")
    print("pip install pandas numpy scipy openpyxl imsciences fredapi cif requests_cache geopy")
    sys.exit(1)

ims = dataprocessing()

def curve_creation(curve_df):
    def objective_abc(x, a, b, c):
        return (a / (1 + b * x**c))

    def objective_abc2(x, a, b):
        return (a / (1 + b * (x**-1)))

    output_data = []

    for column in curve_df.columns:
        if column == "Spend":
            continue

        current_data = curve_df.dropna(subset=["Spend", column])
        # Ensure we have at least some data
        if current_data.empty or len(current_data) < 3:
            print(f"{column} does not have enough data points, skipping.")
            continue

        x = pd.to_numeric(current_data["Spend"].astype(str).str.replace(",", ""), errors='coerce').dropna()
        y = pd.to_numeric(current_data[column].astype(str).str.replace(",", ""), errors='coerce').dropna()

        # Re-check after dropna
        if x.empty or y.empty or len(x) != len(y) or len(x) < 3:
            print(f"{column} does not have sufficient clean data, skipping.")
            continue

        if y.sum() == 0:
            print(f"{column} no data (all zeros), skipping.")
            continue

        # Initial guesses
        mid_index = len(x) // 2
        a_guess = float(y.iloc[mid_index] * 2)
        b_guess = float(x.iloc[mid_index])

        # Ensure a_guess and b_guess are finite
        if not np.isfinite(a_guess) or not np.isfinite(b_guess):
            print(f"{column} initial guesses are not finite, skipping.")
            continue

        attempt_params = [
            {"bounds": ([0, 0, -np.inf], [np.inf, np.inf, 0]), "p0": [a_guess, b_guess, -1]},
            {"bounds": ([0, 0], [np.inf, np.inf]), "p0": [a_guess, b_guess]},
            {"bounds": ([0, 0, -1.5], [np.inf, np.inf, -0.5]), "p0": [a_guess, b_guess, -1]}
        ]

        popt = None
        for params in attempt_params:
            try:
                if len(params['p0']) == 3:
                    popt, _ = optimize.curve_fit(objective_abc, x, y, **params)
                else:
                    popt, _ = optimize.curve_fit(objective_abc2, x, y, **params)
                # If successful, break out of the loop
                break
            except (RuntimeError, ValueError):
                # Try next attempt parameters
                continue

        if popt is None:
            # If still None, try a gradient-based fallback or skip
            df_tester = pd.concat([x, y], axis=1)
            df_tester.columns = ["Spend", "Value"]
            df_tester["Gradient"] = df_tester["Value"].diff() / df_tester["Spend"].diff()
            df_tester.index = np.arange(len(df_tester))
            
            # Check if we can find a minimum gradient point
            if df_tester["Gradient"].iloc[1:].isnull().all():
                # If gradient calculation failed or we have no valid data
                print(f"Failed to fit curve for {column} - no gradient approach possible.")
                continue
            
            minimum_idx = df_tester["Gradient"].iloc[1:].idxmin()
            df_tester_2 = df_tester.iloc[:minimum_idx]
            
            x = pd.to_numeric(df_tester_2["Spend"])
            y = pd.to_numeric(df_tester_2["Value"])
            
            if x.empty or y.empty or len(x) < 2:
                print(f"Failed to fit curve for {column} - insufficient data after gradient approach.")
                continue

            try:
                popt, _ = optimize.curve_fit(objective_abc2, x, y,
                                             bounds=([0, 0], [np.inf, np.inf]),
                                             p0=[a_guess, b_guess])
            except (RuntimeError, ValueError):
                print(f"Failed to fit curve for {column}")
                continue

        # Compute final parameters and R^2
        if len(popt) == 3:
            a, b, c = popt
            epsilon = 1e-6
            b_normal = b ** (-1 / c) if abs(c) > epsilon else np.nan
            y_pred = objective_abc(x, a, b, c)
        else:
            a, b = popt
            c = np.nan
            b_normal = np.nan
            y_pred = objective_abc2(x, a, b)

        residuals = y - y_pred
        ss_res = np.sum(residuals**2)
        ss_tot = np.sum((y - np.mean(y))**2)
        if ss_tot == 0:
            # Avoid division by zero if there's no variance in y
            r_squared = np.nan
        else:
            r_squared = 1 - (ss_res / ss_tot)

        output_data.append({
            'Curve Name': column,
            'A': a,
            'B': b,
            'C': c,
            'B_normal': b_normal,
            'R^2': r_squared,
            'Type': curve_df[column].iloc[0]
        })

    return pd.DataFrame(output_data)

def calculate_spend_curve(df, spend_increment=10000, max_spend=10010000):
    """
    Calculate spend curve values for each row in the given DataFrame.

    Parameters:
    df (pd.DataFrame): The DataFrame containing columns 'Curve Name', 'A', 'B', 'C'.
    spend_increment (int): The increment value for spend.
    max_spend (int): The maximum spend value to calculate.

    Returns:
    pd.DataFrame: A new DataFrame with calculated spend curve values in a wide format and parameters.
    """
    spend_values = np.arange(0, max_spend, spend_increment)
    results = {}
    warnings.filterwarnings("ignore", category=RuntimeWarning)

    parameter_dict = {}
    for _, row in df.iterrows():
        A = row['A']
        B = row['B']
        C = row['C']
        curve_name = row['Curve Name']
        
        values = []
        for spend in spend_values:
            if spend == 0:
                value = 0
            else:
                try:
                    value = A / (1 + B * (spend ** C))
                except ZeroDivisionError:
                    value = np.nan
            values.append(value)

        results[curve_name] = values
        parameter_dict[curve_name] = {'A': A, 'B': B, 'C': C}

    results_df = pd.DataFrame(results, index=spend_values)
    results_df.index.name = 'Spend'
    return results_df, parameter_dict

def generate_abc_curves(import_file, output_file_path, skip_rows=0, spend_increment=10000, max_spend=10010000):
    """
    Process response curves from an Excel file and generate an output with calculated ABC curves.

    Args:
        import_file (str): Path to the input Excel file.
        output_file_path (str): Path to save the output Excel file (should be .xlsx).
        skip_rows (int): Number of rows to skip when reading the input sheet.
        spend_increment (int): The increment value for spend.
        max_spend (int): The maximum spend value to calculate.
    """
    if not os.path.exists(import_file):
        raise FileNotFoundError(f"The file {import_file} does not exist.")
    
    # Load the data
    df_response_curves = pd.read_excel(import_file, skiprows=skip_rows)
    df_response_curves = df_response_curves.iloc[:, 6:]
    df_response_curves = df_response_curves.dropna(subset=[df_response_curves.columns[1]])
    if 'Current' in df_response_curves.columns:
        df_response_curves = df_response_curves.drop(columns=['Current'])

    # Filter out columns where numeric sum is 0
    numeric_cols = df_response_curves.select_dtypes(include=[np.number]).columns
    non_zero_cols = numeric_cols[df_response_curves[numeric_cols].sum() != 0]
    non_numeric_cols = df_response_curves.select_dtypes(exclude=[np.number]).columns
    df_response_curves = df_response_curves[list(non_numeric_cols) + list(non_zero_cols)]
    print(df_response_curves)
    # Generate curve parameters
    curve_equation_parameters = curve_creation(df_response_curves)

    # Calculate spend curves
    results_df, parameter_dict = calculate_spend_curve(curve_equation_parameters, spend_increment, max_spend)
    results_df.reset_index(drop=False, inplace=True)

    # Ensure output_file_path ends with .xlsx
    if not output_file_path.lower().endswith('.xlsx'):
        raise ValueError("output_file_path must end with '.xlsx'")

    if not os.path.exists(output_file_path):
        # Create a new file with parameters and results
        with pd.ExcelWriter(output_file_path, mode='w', engine='openpyxl') as writer:
            # Write ABC parameters as numeric values
            params_data = pd.DataFrame({
                curve_name: [params['A'], params['B'], params['C']]
                for curve_name, params in parameter_dict.items()
            }, index=['Parameter A', 'Parameter B', 'Parameter C'])
            params_data.to_excel(writer, sheet_name='ABC Curves', index=True, header=True, startrow=0)

            # Write the results DataFrame
            results_df.to_excel(writer, sheet_name='ABC Curves', index=False, startrow=len(params_data) + 2)

    else:
        # Load the existing workbook
        book = load_workbook(output_file_path)

        # Remove existing sheet if it exists
        if 'ABC Curves' in book.sheetnames:
            del book['ABC Curves']

        # Create a new sheet
        sheet = book.create_sheet(title='ABC Curves')

        # Write parameter headers and values
        sheet.append([''] + list(parameter_dict.keys()))
        sheet.append(['Parameter A'] + [param['A'] for param in parameter_dict.values()])
        sheet.append(['Parameter B'] + [param['B'] for param in parameter_dict.values()])
        sheet.append(['Parameter C'] + [param['C'] for param in parameter_dict.values()])

        # Empty row for separation
        sheet.append([])

        # Add column names
        sheet.append(['Spend'] + list(parameter_dict.keys()))

        # Populate spend values and formulas
        start_row = 7 
        for i, spend_value in enumerate(np.arange(0, max_spend, spend_increment)):
            # Write spend value
            sheet.cell(row=start_row + i, column=1, value=spend_value)
            for col_idx, curve_name in enumerate(parameter_dict.keys(), start=2):
                A_cell = sheet.cell(row=2, column=col_idx).coordinate
                B_cell = sheet.cell(row=3, column=col_idx).coordinate
                C_cell = sheet.cell(row=4, column=col_idx).coordinate
                if spend_value == 0:
                    formula = "=0"
                else:
                    formula = f"={A_cell}/(1+{B_cell}*({sheet.cell(row=start_row+i, column=1).coordinate}^{C_cell}))"
                sheet.cell(row=start_row + i, column=col_idx, value=formula)

        # Save the workbook
        book.save(output_file_path)

# MAIN EXECUTION:
import_file = os.path.join("C:\\Users\\CameronRoberts\\OneDrive - im-sciences.com\\Documents\\01-2025", "ABC_Input", "Response_Curves.xlsx")
output_file_path = os.path.join("C:\\Users\\CameronRoberts\\OneDrive - im-sciences.com\\Documents\\01-2025", "ABC_Output", "ABC_Curves_Output.xlsx")

# Configurable parameters
spend_increment = 10000
max_spend = 20010000

# Generate the ABC curves
generate_abc_curves(import_file, output_file_path, spend_increment=spend_increment, max_spend=max_spend)

